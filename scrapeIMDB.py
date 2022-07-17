# Importing the modules:
from bs4 import BeautifulSoup
import requests
import openpyxl

# Creating the excel file for the records:
excel = openpyxl.Workbook()
# print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Top 250 Movies - IMDB'
print(excel.sheetnames)
sheet.append(['Movie Rank', 'Movie Name', 'Year of Release', 'IMDB Rating'])

# We want to extract or scrape the IMDB website for information of Top 250 Movies:
# 1] Name of the movie
# 2] Ranking of the movie
# 3] Year it was released
# 4] IMDB Rating

try:
    source = requests.get("https://www.imdb.com/chart/top/?ref_=nv_mv_250")
    source.raise_for_status()

    soup = BeautifulSoup(source.text, 'html.parser')
    #print(soup)

    # accessing the <tbody> tag for getting all the information:
    movies = soup.find('tbody', class_ = 'lister-list').find_all('tr')
    #print(len(movies))
    for movie in movies:
        
        name = movie.find('td', class_ ='titleColumn').a.text 

        rank = movie.find('td', class_='titleColumn').get_text(strip=True).split('.')[0]       

        year = movie.find('td',class_='titleColumn').span.text.strip("()")

        rating = movie.find('td', class_='ratingColumn imdbRating').strong.text
        
        print(rank, name, year, rating)
        sheet.append([rank, name, year, rating])
        


except Exception as e:
    print(e)

# SAVE THE EXCEL FILE:
excel.save('IMDB Movie Ratings.xlsx')